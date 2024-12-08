from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from copy import copy


class AdmissionDiffChecker:
    def __init__(self):
        self.old_admission_map = {}
        self.header_row = None
        self.diff_rows = []
        self.same_rows = []

    def load_older_worksheet(self):
        print("Parsing Older Worksheet")
        wb = load_workbook(
            filename="assets/Bewertungen_2024-11-05_old.xlsx"
        )  # adjust filename for each usecase
        ws = wb.active
        for row in ws.iter_rows(
            min_row=2, max_row=91, values_only=False
        ):  # adjust max row depending on number of entries of relevant semester
            if not row[4].value or not row[5].value:
                continue
            student_name = row[4].value + row[5].value
            # We just copy the current style applied to the L column (admission status)
            self.old_admission_map[student_name] = copy(row[11].fill)

    def load_and_compare_newer_worksheet(self):
        print("Parsing Newer Worksheet")
        wb = load_workbook(
            filename="assets/Bewertungen_2024-11-05_new.xlsx"
        )  # adjust filename for each usecase
        ws = wb.active
        for row in ws.iter_rows(
            min_row=1, max_row=91, values_only=True
        ):  # adjust max row depending on number of entries of relevant semester
            # Write header row for result
            if not self.header_row:
                self.header_row = row
                continue
            if not row[4] or not row[5]:
                continue
            name = row[4] + row[5]
            if name not in self.old_admission_map:
                self.diff_rows.append(row)
                # This is for de-duplication, add a default white colour background
                self.old_admission_map[name] = None
            # Save old entries for creating excl with old+new entries
            else:
                self.same_rows.append(row)

    def write_result_to_new_worksheet(self):

        # Excl of difference only
        wb = Workbook()
        ws = wb.active
        # Sort by surname
        self.diff_rows.sort(key=lambda row: row[4].lower())
        self.same_rows.sort(key=lambda row: row[4].lower())
        ws.append(self.header_row)
        for row in self.diff_rows:
            ws.append(row)
        wb.save("assets/only_diff.xlsx")

        # Excl of all entries
        wb2 = Workbook()
        ws = wb2.active
        ws.append(self.header_row)
        for row in self.diff_rows:
            ws.append(row)
        ws.append([])

        same_rows_start = len(self.diff_rows) + 3
        for i, row in enumerate(self.same_rows):
            ws.append(row)
            name = row[4] + row[5]
            # Apply appropriate styling
            if self.old_admission_map[name]:
                ws[f"L{same_rows_start + i}"].fill = self.old_admission_map[name]
        wb2.save("assets/all_entries_sorted.xlsx")
        """
        unsorted = pd.read_excel("assets/all_entries1.xlsx")
        print(unsorted)
        unsorted.sort_values(by="BewName", ascending=True)
        unsorted.to_excel("assets/all_entries_sorted.xlsx")
        """

    def do_comparison(self):
        self.load_older_worksheet()
        self.load_and_compare_newer_worksheet()
        self.write_result_to_new_worksheet()


start_time = datetime.now()
print(f"Begun Checking Diffs {start_time}")
checker = AdmissionDiffChecker()
checker.do_comparison()
finish_time = datetime.now() - start_time
print(f"Finished Checking Diffs. Time Taken: {finish_time.microseconds} microseconds")
